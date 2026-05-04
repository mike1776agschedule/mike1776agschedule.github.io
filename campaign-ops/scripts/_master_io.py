"""
Shared helper for surgical updates to the master campaign workbook.

The master is the boss's `florida_sign_operations_master_*.xlsx` file.
We classify its sheets:

  BOSS_OWNED     — never modified except via update_cells() for the
                   feedback-loop columns of County_Master
  USER_EDITED    — empty operational sheets the user fills in by hand;
                   never overwritten
  DERIVED        — sheets we add (Dashboard, Directory_Audit) which we
                   regenerate on every refresh

This module never touches sheets it isn't asked to touch.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

DEFAULT_MASTER_PATH = "outputs/campaign_ops_master.xlsx"

# Boss-owned sheets in the canonical master (treated as read-only except via update_cells).
BOSS_OWNED_SHEETS = [
    "Overview",
    "County_Master",
    "Partners",
    "County_Source_Log",
    "Influence_Prospects",
    "Political_Committee_Leads",
    "County_Audit",
    "Top_Target_Counties",
    "Priority_Call_List",
    "Regional_Hubs",
    "Sign_Site_Playbook",
    "Outdoor_Advertising_Leads",
    "Audit_1_Data",
    "Audit_2_Operational",
]

# Operational sheets the user fills in. These existed empty in the boss's file.
USER_EDITED_SHEETS = [
    "Yard_Sign_Deliveries",
    "Large_Sign_Locations",
    "Outreach_Log",
]

# Derived sheets we add and refresh.
DERIVED_SHEETS = [
    "Executive_Summary",
    "Dashboard",
    "Directory_Audit",
    "Sign_Location_Candidates",
    "Sign_Deployment_Plan",
    "Yard_Sign_Allocation",
    "QC_Report",
]


def read_settings(path: str | Path) -> dict:
    """Read the Settings sheet and return a dict of input values.

    Falls back to defaults if Settings sheet is missing or values are unparseable.
    """
    defaults = {
        "TOTAL_4X8_PRIMARY":  2000,
        "TOTAL_YARD_SIGNS":   4000,
        "REPLACEMENT_RATIO":  0.25,
        "MIN_PER_COUNTY_4X8": 4,
        "TIER_A_FLOOR":       100,
        "TIER_B_FLOOR":       60,
        "TIER_C_FLOOR":       40,
        "TIER_D_FLOOR":       25,
    }
    p = Path(path)
    if not p.exists():
        return defaults
    from openpyxl import load_workbook
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception:
        return defaults
    if "Settings" not in wb.sheetnames:
        return defaults
    ws = wb["Settings"]
    out = dict(defaults)
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None or len(row) < 2:
            continue
        name = str(row[0]).strip()
        if name in defaults:
            v = row[1]
            if v is None:
                continue
            try:
                if isinstance(defaults[name], int):
                    out[name] = int(float(v))
                elif isinstance(defaults[name], float):
                    out[name] = float(v)
                else:
                    out[name] = v
            except (TypeError, ValueError):
                pass
    return out


def list_sheets(path: str | Path) -> list[str]:
    from openpyxl import load_workbook

    p = Path(path)
    if not p.exists():
        return []
    return load_workbook(p, read_only=True).sheetnames


def _load_and_freeze(path: Path):
    """
    Load the workbook with cached formula values, then materialize formulas
    as static values so subsequent pandas reads always return the correct
    numbers. The boss's file contains cross-sheet formulas (e.g.
    Top_Target_Counties references County_Master); openpyxl preserves the
    formula but loses Excel's cached value on save, which would break our
    pandas reads. Freezing once on first save makes the pipeline reliable.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        # iter_rows here returns Cell objects; reassigning .value to itself
        # converts any formula cell to its cached value (data_only mode).
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.value = cell.value
    return wb


def replace_sheet(
    path: str | Path,
    sheet_name: str,
    df,  # pandas.DataFrame
    *,
    autofilter: bool = True,
    freeze_top: bool = True,
    color_col: str | None = None,
    color_map: dict | None = None,
    dropdowns: dict | None = None,  # {column_name: [allowed values]}
) -> None:
    """
    Replace a single sheet with `df` contents. Other sheets are untouched.

    Optional styling:
      - autofilter:  adds an autofilter over the data range
      - freeze_top:  freezes row 1 (the header)
      - color_col:   name of a column whose value selects a row tint
      - color_map:   dict of {value: hex-rgb-without-#} for color_col
    """
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"master workbook missing: {p}")
    wb = _load_and_freeze(p)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    if df is None or len(df.columns) == 0:
        wb.save(p)
        return

    headers = list(df.columns)
    ws.append(headers)
    for row in df.itertuples(index=False, name=None):
        ws.append([_xlsx_safe(v) for v in row])

    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    n_rows = len(df) + 1
    n_cols = len(headers)
    if autofilter and n_rows > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
    if freeze_top:
        ws.freeze_panes = "A2"

    # Excel data validation dropdowns
    if dropdowns:
        from openpyxl.worksheet.datavalidation import DataValidation
        for col_name, allowed in dropdowns.items():
            if col_name not in headers:
                continue
            col_idx = headers.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            # Excel limits formula1 to 255 chars; comma-join values
            formula = '"' + ",".join(str(v).replace('"', "'") for v in allowed) + '"'
            if len(formula) > 255:
                # Too long for inline — skip dropdown but keep coloring
                continue
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.error = f"Pick one of: {', '.join(allowed)}"
            dv.errorTitle = "Invalid status"
            dv.add(f"{col_letter}2:{col_letter}{n_rows}")
            ws.add_data_validation(dv)

    if color_col and color_map and color_col in headers:
        ccol = headers.index(color_col) + 1
        fills = {k: PatternFill("solid", fgColor=v) for k, v in color_map.items()}
        for row_idx in range(2, n_rows + 1):
            v = ws.cell(row=row_idx, column=ccol).value
            fill = fills.get(v)
            if fill:
                for c in range(1, n_cols + 1):
                    ws.cell(row=row_idx, column=c).fill = fill

    wb.save(p)


def read_sheet_safe(path: str | Path, sheet_name: str):
    """Return the sheet as a pandas DataFrame; empty df if sheet missing or empty."""
    import pandas as pd

    p = Path(path)
    if not p.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(p, sheet_name=sheet_name)
    except ValueError:
        return pd.DataFrame()


def update_cells(
    path: str | Path,
    sheet_name: str,
    key_col: str,
    updates: dict[str, dict[str, object]],
) -> int:
    """
    Surgically update specific cells in `sheet_name`. Used by the feedback
    loop to write `Yard Signs Delivered` / `4x8 Confirmed` into County_Master
    without touching any other column.

    Args:
      path:       master workbook
      sheet_name: target sheet (e.g. 'County_Master')
      key_col:    column whose value identifies a row (e.g. 'County')
      updates:    {key_value: {column_name: new_value}}, e.g.
                  {'Alachua': {'Yard Signs Delivered': 25, 'Yard Signs Remaining': 50}}

    Returns the number of rows updated.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"master workbook missing: {p}")
    wb = _load_and_freeze(p)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet not in workbook: {sheet_name}")
    ws = wb[sheet_name]

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if key_col not in headers:
        raise ValueError(f"key column {key_col!r} not found in {sheet_name} headers")
    key_idx = headers.index(key_col) + 1
    col_idx = {h: headers.index(h) + 1 for h in headers}

    rows_changed = 0
    for row in ws.iter_rows(min_row=2):
        key_val = row[key_idx - 1].value
        if key_val is None:
            continue
        key_str = str(key_val).strip()
        if key_str not in updates:
            continue
        for col_name, new_val in updates[key_str].items():
            if col_name not in col_idx:
                continue
            ws.cell(row=row[0].row, column=col_idx[col_name], value=_xlsx_safe(new_val))
        rows_changed += 1

    wb.save(p)
    return rows_changed


def _xlsx_safe(v):
    """openpyxl rejects pandas NaN, NaT, and a few other types; coerce to None/str."""
    try:
        import pandas as pd
        if v is pd.NaT:
            return None
        if isinstance(v, float):
            import math
            if math.isnan(v):
                return None
        if hasattr(v, "isoformat") and not isinstance(v, str):
            return v
        if isinstance(v, (list, tuple, dict, set)):
            return str(v)
        return v
    except Exception:
        return v
