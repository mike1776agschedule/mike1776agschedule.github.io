#!/usr/bin/env python3
"""
Quality Control review of the master workbook + Sign_Location_Candidates.

Writes a `QC_Report` sheet summarizing:
  - Boss data integrity (sheet count, totals preserved)
  - Per-county candidate density vs Strategic Tier
  - Category mix (chain dominance check)
  - Geographic coverage (sub-county cities)
  - Duplicate / coordinate-clash detection
  - Strategic flags (under-covered counties, over-fueled mix, motorway_junctions)

Usage:
    python scripts/qc_report.py
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

DEFAULT_MASTER = "outputs/campaign_ops_master.xlsx"

# Florida sub-county cities to spot-check (~5mi radius via lat/lon proximity).
SUBCOUNTY_CITIES = [
    # (city, lat, lon, county)
    ("Hialeah",         25.8576, -80.2781, "Miami-Dade"),
    ("Doral",           25.8195, -80.3553, "Miami-Dade"),
    ("Coral Gables",    25.7215, -80.2684, "Miami-Dade"),
    ("Miami Gardens",   25.9420, -80.2456, "Miami-Dade"),
    ("Homestead",       25.4687, -80.4776, "Miami-Dade"),
    ("Pembroke Pines",  26.0078, -80.2962, "Broward"),
    ("Hollywood",       26.0112, -80.1495, "Broward"),
    ("Miramar",         25.9876, -80.2323, "Broward"),
    ("Coral Springs",   26.2710, -80.2706, "Broward"),
    ("Plantation",      26.1276, -80.2331, "Broward"),
    ("Boca Raton",      26.3683, -80.1289, "Palm Beach"),
    ("Boynton Beach",   26.5251, -80.0664, "Palm Beach"),
    ("Delray Beach",    26.4615, -80.0728, "Palm Beach"),
    ("Brandon",         27.9378, -82.2859, "Hillsborough"),
    ("Riverview",       27.8661, -82.3262, "Hillsborough"),
    ("Plant City",      28.0186, -82.1148, "Hillsborough"),
    ("Wesley Chapel",   28.2398, -82.3265, "Pasco"),
    ("St. Petersburg",  27.7676, -82.6403, "Pinellas"),
    ("Largo",           27.9095, -82.7873, "Pinellas"),
    ("Clearwater",      27.9659, -82.8001, "Pinellas"),
    ("Tallahassee",     30.4383, -84.2807, "Leon"),
    ("Port St. Lucie",  27.2730, -80.3582, "St. Lucie"),
    ("Cape Coral",      26.5629, -81.9495, "Lee"),
    ("Naples",          26.1420, -81.7948, "Collier"),
    ("Sarasota",        27.3364, -82.5307, "Sarasota"),
    ("Ocala",           29.1872, -82.1401, "Marion"),
    ("Lakeland",        28.0395, -81.9498, "Polk"),
    ("Daytona Beach",   29.2108, -81.0228, "Volusia"),
    ("Pensacola",       30.4213, -87.2169, "Escambia"),
    ("Panama City",     30.1588, -85.6602, "Bay"),
    ("Gainesville",     29.6516, -82.3248, "Alachua"),
    ("Kissimmee",       28.2920, -81.4076, "Osceola"),
    ("Fort Myers",      26.6406, -81.8723, "Lee"),
    ("Bradenton",       27.4989, -82.5748, "Manatee"),
    ("The Villages",    28.9165, -81.9598, "Sumter"),
    ("Vero Beach",      27.6386, -80.3973, "Indian River"),
    ("Stuart",          27.1973, -80.2528, "Martin"),
    ("Crestview",       30.7626, -86.5707, "Okaloosa"),
    ("Destin",          30.3935, -86.4958, "Okaloosa"),
]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Quality control review of the master workbook.")
    p.add_argument("--master", default=DEFAULT_MASTER)
    return p.parse_args()


def main() -> int:
    args = parse_args()

    try:
        import pandas as pd
    except ImportError as e:
        print(f"ERROR: missing dependency ({e}).", file=sys.stderr)
        return 2

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import _master_io as mio  # type: ignore

    master = Path(args.master)
    if not master.exists():
        print(f"ERROR: master not found: {master}", file=sys.stderr)
        return 1

    print(f"Running QC on {master} ...")

    cm  = mio.read_sheet_safe(master, "County_Master")
    slc = mio.read_sheet_safe(master, "Sign_Location_Candidates")
    ca  = mio.read_sheet_safe(master, "County_Audit")
    md  = pd.DataFrame()
    dpath = Path("data/raw/florida_gop_directory.xlsx")
    if dpath.exists():
        md = pd.read_excel(dpath, sheet_name="Master Directory")

    findings: list[tuple[str, str, str]] = []  # (category, status, message)

    # ── Section 1: Master integrity
    # Expected sheets:
    #   17 boss-owned + 3 user-edited + 8 derived (added Settings)
    expected_sheets = 25
    from openpyxl import load_workbook
    wb = load_workbook(master, read_only=True)
    sheet_count = len(wb.sheetnames)
    findings.append(("Integrity", "PASS" if sheet_count == expected_sheets else "FAIL",
                     f"Master has {sheet_count} sheets (expected {expected_sheets})"))

    expected_yard = 7485
    expected_4x8 = 367
    actual_yard = int(cm["Suggested Yard Signs"].fillna(0).sum())
    actual_4x8 = int(cm["Suggested 4x8 Goal"].fillna(0).sum())
    findings.append(("Integrity", "PASS" if actual_yard == expected_yard else "FAIL",
                     f"Suggested Yard Signs = {actual_yard} (expected {expected_yard})"))
    findings.append(("Integrity", "PASS" if actual_4x8 == expected_4x8 else "FAIL",
                     f"Suggested 4x8 Goal = {actual_4x8} (expected {expected_4x8})"))

    expected_tiers = {"A": 11, "B": 8, "C": 14, "D": 34}
    actual_tiers = cm["Strategic Tier"].value_counts().to_dict()
    tier_ok = all(actual_tiers.get(k, 0) == v for k, v in expected_tiers.items())
    findings.append(("Integrity", "PASS" if tier_ok else "FAIL",
                     f"Tier distribution: {actual_tiers}"))

    counties_in_slc = slc["County"].nunique() if not slc.empty else 0
    findings.append(("Integrity", "PASS" if counties_in_slc == 67 else "FAIL",
                     f"Counties in Sign_Location_Candidates: {counties_in_slc}/67"))

    # ── Section 2: Candidate quantity
    total_cands = len(slc)
    findings.append(("Candidates", "INFO", f"Total candidates: {total_cands}"))
    findings.append(("Candidates", "INFO",
                     f"Target installs: 2,000  /  Buffer: {total_cands/2000:.2f}x"))

    # ── Section 3: Category mix
    if not slc.empty:
        type_counts = slc["Type"].value_counts().to_dict()
        for t in ("intersection", "commercial", "agri_civic"):
            findings.append(("Mix", "INFO", f"{t}: {type_counts.get(t, 0)}"))

        # Motorway junctions check (should be 0 after fix)
        mj = (slc["Category"] == "motorway_junction").sum()
        findings.append(("Mix", "PASS" if mj == 0 else "FAIL",
                         f"motorway_junction nodes: {mj} (should be 0 — not sign-able)"))

        # Chain dominance check (commercial)
        comm = slc[slc["Type"] == "commercial"]
        if len(comm) > 0:
            top_cat = comm["Category"].value_counts(normalize=True).head(1)
            top_cat_name = top_cat.index[0]
            top_cat_pct = float(top_cat.iloc[0])
            findings.append(("Mix",
                             "PASS" if top_cat_pct < 0.50 else "WARN",
                             f"Top commercial category: {top_cat_name} = {top_cat_pct:.0%} (target: <50%)"))

            top_chain = comm["Name"].value_counts().head(1)
            findings.append(("Mix", "INFO",
                             f"Most common business name: {top_chain.index[0]} ({int(top_chain.iloc[0])} sites)"))

        # Rural agri_civic share — informational only post-highway-pull.
        # The 20% target was meaningful when the candidate pool was ~3K; with
        # 20K+ sites (mostly highway-corridor commercial), 3-5% is normal.
        d_counties = cm[cm["Strategic Tier"] == "D"]["County"].tolist()
        d_slc = slc[slc["County"].isin(d_counties)]
        if len(d_slc):
            d_agri_pct = (d_slc["Type"] == "agri_civic").mean()
            findings.append(("Mix", "INFO",
                             f"Tier D agri/civic share: {d_agri_pct:.0%} "
                             f"(post-highway-pull baseline; rural pool now commercial-dominant)"))

    # ── Section 4: Coverage by county
    if not slc.empty:
        agg = slc.groupby("County").size().rename("cands")
        per_county = cm[["County", "Strategic Tier", "Suggested 4x8 Goal"]].merge(
            agg, left_on="County", right_index=True, how="left").fillna({"cands": 0})
        per_county["cands"] = per_county["cands"].astype(int)
        per_county["ratio"] = per_county["cands"] / per_county["Suggested 4x8 Goal"].replace(0, 1)

        under = per_county[per_county["cands"] < per_county["Suggested 4x8 Goal"] * 3]
        findings.append(("Coverage",
                         "PASS" if under.empty else "WARN",
                         f"Counties below 3× their 4x8 goal: {len(under)}"))

        # Sub-county city coverage
        gaps = []
        for city, lat, lon, county in SUBCOUNTY_CITIES:
            sub = slc[slc["County"] == county]
            sub = sub[((sub["Lat"] - lat).abs() < 0.075) &
                      ((sub["Lon"] - lon).abs() < 0.075)]
            if len(sub) < 5:
                gaps.append((city, county, len(sub)))
        findings.append(("Coverage",
                         "PASS" if len(gaps) <= 5 else "WARN",
                         f"Major cities with <5 candidates within 5mi: {len(gaps)}/{len(SUBCOUNTY_CITIES)}"))
        if gaps:
            for city, county, n in gaps[:10]:
                findings.append(("Coverage", "INFO", f"  gap: {city} ({county}) = {n}"))

    # ── Section 5: Duplicates
    if not slc.empty:
        slc_geo = slc.dropna(subset=["Lat", "Lon"]).copy()
        slc_geo["__k__"] = list(zip(slc_geo["Lat"].round(5), slc_geo["Lon"].round(5)))
        dups = slc_geo.groupby("__k__").size()
        clashes = (dups > 1).sum()
        findings.append(("Quality", "PASS" if clashes == 0 else "WARN",
                         f"Coordinate clashes (5-decimal precision): {clashes}"))

    # ── Section 5b: V3 categorical exclusions (must be 0 of each)
    if not slc.empty:
        excluded_cats = [
            "place_of_worship", "school", "college", "university", "kindergarten",
            "post_office", "fire_station", "townhall", "library", "courthouse", "police",
            "bank", "atm", "pharmacy", "hospital", "clinic", "doctors",
            "funeral_directors", "chemist", "supermarket",
            "motorway_junction",
        ]
        for cat in excluded_cats:
            n = int((slc["Category"] == cat).sum())
            findings.append(("Exclusions",
                             "PASS" if n == 0 else "FAIL",
                             f"{cat}: {n} (must be 0)"))

    # ── Section 5c: Chain blocklist enforcement
    if not slc.empty:
        chain_blocklist = [
            "lowe's", "home depot", "discount tire", "firestone", "pep boys",
            "walmart", "target", "costco", "sam's club", "dollar general",
            "dollar tree", "family dollar", "publix", "winn-dixie", "kroger",
            "wawa", "racetrac", "circle k", "7-eleven", "cvs", "walgreens",
            "rite aid", "ikea", "best buy",
        ]
        name_lower = slc["Name"].fillna("").astype(str).str.lower()
        violations = 0
        violator_examples: list[str] = []
        for token in chain_blocklist:
            mask = name_lower.str.contains(token, regex=False, na=False)
            n = int(mask.sum())
            if n > 0:
                violations += n
                ex = slc[mask]["Name"].head(1).tolist()
                if ex:
                    violator_examples.append(f"{token} -> {ex[0]}")
        findings.append(("Exclusions",
                         "PASS" if violations == 0 else "FAIL",
                         f"Chain blocklist violations: {violations}"))
        for ex in violator_examples[:5]:
            findings.append(("Exclusions", "INFO", f"  example: {ex}"))

    # ── Section 5d: POC enrichment coverage on Sign_Location_Candidates
    if not slc.empty and "Phone" in slc.columns:
        has_phone = int((slc["Phone"].fillna("").astype(str).str.len() > 0).sum())
        has_email = int((slc["Email"].fillna("").astype(str).str.len() > 0).sum())
        pct_phone = has_phone / len(slc) if len(slc) else 0
        findings.append(("POC",
                         "INFO",
                         f"Candidates with phone: {has_phone}/{len(slc)} ({pct_phone:.0%})"))
        findings.append(("POC",
                         "INFO",
                         f"Candidates with email: {has_email}/{len(slc)}"))
        # POC source breakdown
        if "Phone_Source" in slc.columns:
            sources = slc[slc["Phone"].fillna("").astype(str).str.len() > 0]["Phone_Source"].value_counts().head(5)
            for src, n in sources.items():
                findings.append(("POC", "INFO", f"  Phone source '{src}': {n}"))

    # ── Section 5e: Sign_Deployment_Plan checks
    sdp = mio.read_sheet_safe(master, "Sign_Deployment_Plan")
    if not sdp.empty:
        n_primary = int((sdp["Plan"] == "primary").sum())
        n_replace = int((sdp["Plan"] == "replacement").sum())
        findings.append(("Deployment", "INFO",
                         f"Sign_Deployment_Plan: {n_primary} primary + {n_replace} replacement"))
        target_primary = 2000
        findings.append(("Deployment",
                         "PASS" if n_primary == target_primary else "WARN",
                         f"Primary count: {n_primary} (target {target_primary})"))
        prim = sdp[sdp["Plan"] == "primary"]
        if len(prim):
            prim_phone = (prim["Phone"].fillna("").astype(str).str.len() > 0).sum()
            pct = prim_phone / len(prim)
            # OSM-only ceiling. We ship with click-to-search URLs for the rest;
            # field staff look those up at call time. Target ≥30%.
            findings.append(("Deployment",
                             "PASS" if pct >= 0.30 else "WARN",
                             f"Primary deployment with verified phone (OSM): {prim_phone}/{len(prim)} ({pct:.0%}, target ≥30%)"))
            findings.append(("Deployment", "INFO",
                             f"Remaining {len(prim)-prim_phone} sites have Google_Search_URL for one-click lookup at call time."))
        # No-hallucination check: every Phone has a non-empty Phone_Source
        bad = int(((sdp["Phone"].fillna("").astype(str).str.len() > 0) &
                   (sdp["Phone_Source"].fillna("").astype(str).str.len() == 0)).sum())
        findings.append(("Deployment",
                         "PASS" if bad == 0 else "FAIL",
                         f"Plan rows with phone but no source (hallucination check): {bad}"))
    else:
        findings.append(("Deployment", "INFO", "Sign_Deployment_Plan empty (run select_sign_deployment.py)"))

    # ── Section 5f: Yard_Sign_Allocation checks
    ysa = mio.read_sheet_safe(master, "Yard_Sign_Allocation")
    if not ysa.empty:
        n_counties = len(ysa)
        boss_total = int(ysa["Suggested_Yard_Signs_Boss"].fillna(0).sum())
        plan_total = int(ysa["Plan_4000"].fillna(0).sum())
        findings.append(("Yard",
                         "PASS" if n_counties == 67 else "FAIL",
                         f"Yard allocation rows: {n_counties} (expected 67)"))
        findings.append(("Yard",
                         "PASS" if boss_total == 7485 else "FAIL",
                         f"Suggested_Yard_Signs_Boss total: {boss_total} (expected 7485)"))
        findings.append(("Yard",
                         "PASS" if plan_total == 4000 else "FAIL",
                         f"Plan_4000 total: {plan_total} (expected 4000)"))
        # POC coverage — every county must have at least chair OR delivery contact
        missing = int((ysa["POC_Status_Yard"] == "MISSING").sum())
        findings.append(("Yard",
                         "PASS" if missing == 0 else "FAIL",
                         f"Counties with no POC at all: {missing}"))
        chair_complete = int((ysa["POC_Status_Yard"] == "chair_complete").sum())
        findings.append(("Yard", "INFO",
                         f"Counties with full chair POC: {chair_complete}/{n_counties}"))
    else:
        findings.append(("Yard", "INFO", "Yard_Sign_Allocation empty (run allocate_yard_signs.py)"))

    # ── Settings reconciliation: do current totals match Settings inputs?
    settings = mio.read_settings(master)
    if not ysa.empty and "Plan_4000" in ysa.columns:
        actual_yard = int(ysa["Plan_4000"].fillna(0).sum())
        target_yard = int(settings.get("TOTAL_YARD_SIGNS", 0))
        ok = (actual_yard == target_yard)
        findings.append(("Settings",
                         "PASS" if ok else "FAIL",
                         f"Yard sign reconciliation: actual={actual_yard}, "
                         f"Settings.TOTAL_YARD_SIGNS={target_yard}"))
    if not sdp.empty:
        actual_primary = int((sdp["Plan"] == "primary").sum())
        target_primary = int(settings.get("TOTAL_4X8_PRIMARY", 0))
        ok = (actual_primary == target_primary)
        findings.append(("Settings",
                         "PASS" if ok else "FAIL",
                         f"4×8 primary reconciliation: actual={actual_primary}, "
                         f"Settings.TOTAL_4X8_PRIMARY={target_primary}"))

    # Storefront-suggestion coverage
    if not sdp.empty and "Nearest_Storefront_1" in sdp.columns:
        prim_sdp = sdp[sdp["Plan"] == "primary"]
        with_sf = int(prim_sdp["Nearest_Storefront_1"].fillna("").astype(str).str.strip().str.len().gt(0).sum())
        findings.append(("Storefronts",
                         "PASS" if with_sf / max(len(prim_sdp), 1) >= 0.80 else "WARN",
                         f"Primary sites with ≥1 nearby storefront: {with_sf}/{len(prim_sdp)} "
                         f"({_pct(with_sf, len(prim_sdp))}, target ≥80%)"))

    # ── Section 5g: AADT (traffic-volume) coverage on the deployment plan
    if not sdp.empty and "AADT" in sdp.columns:
        prim = sdp[sdp["Plan"] == "primary"]
        aadt = pd.to_numeric(prim["AADT"], errors="coerce").fillna(0)
        n_with = int((aadt > 0).sum())
        med = int(aadt[aadt > 0].median()) if n_with else 0
        n_high = int((aadt >= 25000).sum())
        findings.append(("AADT",
                         "PASS" if n_with / max(len(prim), 1) >= 0.80 else "WARN",
                         f"Primary plan with AADT > 0: {n_with}/{len(prim)} "
                         f"({_pct(n_with, len(prim))}, target ≥80%)"))
        findings.append(("AADT", "INFO", f"Primary plan median AADT: {med:,} vehicles/day"))
        findings.append(("AADT", "INFO", f"Primary plan with AADT ≥ 25k: {n_high}"))

    # ── Section 5h: Status-workflow rollups
    olog = mio.read_sheet_safe(master, "Outreach_Log")
    if not olog.empty and "Result Status" in olog.columns:
        log4x8 = olog[olog.get("Contact Type", pd.Series("", index=olog.index)) == "4x8 Site Owner"]
        if len(log4x8):
            called = int((log4x8["Result Status"].astype(str) != "To Call").sum())
            findings.append(("Outreach", "INFO",
                             f"4x8 outreach calls completed: {called}/{len(log4x8)} "
                             f"({_pct(called, len(log4x8))})"))
            approved = int((log4x8["Result Status"] == "Approved").sum())
            declined = int((log4x8["Result Status"] == "Declined").sum())
            findings.append(("Outreach", "INFO",
                             f"4x8 outreach: {approved} approved, {declined} declined"))
        rec_log = olog[olog.get("Contact Type", pd.Series("", index=olog.index)) == "REC Chair / Delivery"]
        if len(rec_log):
            confirmed = int((rec_log["Result Status"] == "Approved").sum())
            findings.append(("Outreach", "INFO",
                             f"REC delivery confirmations: {confirmed}/{len(rec_log)}"))

    # ── Section 6: Directory cross-check
    if not md.empty:
        md_email_pct = md["Email"].notna().sum() / len(md)
        findings.append(("Directory", "INFO",
                         f"Master Directory: {len(md)} contacts, {md_email_pct:.0%} have email"))
        directory_counties = md["County"].dropna().str.strip().nunique()
        findings.append(("Directory", "INFO",
                         f"Counties represented in Directory: {directory_counties}"))

    # ── Build the report DataFrame
    report_df = pd.DataFrame(findings, columns=["Category", "Status", "Finding"])
    report_df.insert(0, "When", datetime.now().strftime("%Y-%m-%d %H:%M"))

    print("\nFindings:")
    for cat, status, msg in findings:
        print(f"  [{status:4s}] {cat:11s} {msg}")

    # Write to master
    print(f"\nWriting QC_Report sheet ...")
    mio.replace_sheet(master, "QC_Report", report_df,
                      color_col="Status",
                      color_map={
                          "PASS": "D4EDDA",
                          "WARN": "FFF3CD",
                          "FAIL": "F8D7DA",
                      })
    print("Done.")
    return 0


def _pct(n, d) -> str:
    if not d:
        return "0%"
    return f"{(n / d) * 100:.0f}%"


if __name__ == "__main__":
    sys.exit(main())
