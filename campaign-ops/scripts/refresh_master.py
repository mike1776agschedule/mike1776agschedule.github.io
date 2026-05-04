#!/usr/bin/env python3
"""
Refresh the master campaign workbook in place.

What this script does (idempotent — safe to run any time):

  1. FEEDBACK LOOP
     - Sums `Yard_Sign_Deliveries.Qty Delivered` by County → writes
       `County_Master.Yard Signs Delivered` (and recomputes Yard Signs Remaining)
     - Counts `Large_Sign_Locations` rows where Install Status == 'Installed'
       by County → writes `County_Master.4x8 Confirmed`

  2. DIRECTORY_AUDIT sheet (regenerated)
     Cross-walks data/raw/florida_gop_directory.xlsx against the master:
       - Chairs in County_Master that don't appear in Directory
       - Email/phone/address coverage by county and category
       - Recommended next contact actions

  3. DASHBOARD sheet (regenerated)
     Live KPI rollup: sign deployment progress, county tier coverage,
     completeness status, top priority counties needing attention.

What this script NEVER touches:
  - The 17 boss-owned sheets (except the specific feedback-loop columns above)
  - Yard_Sign_Deliveries / Large_Sign_Locations / Outreach_Log (you fill these)

Usage:
    python scripts/refresh_master.py
    python scripts/refresh_master.py --master <path> --directory <path>
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

DEFAULT_MASTER = "outputs/campaign_ops_master.xlsx"
DEFAULT_DIRECTORY = "data/raw/florida_gop_directory.xlsx"


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Refresh the master campaign workbook (Dashboard + Directory_Audit + feedback loop).")
    p.add_argument("--master", default=DEFAULT_MASTER, help=f"Master xlsx (default: {DEFAULT_MASTER})")
    p.add_argument("--directory", default=DEFAULT_DIRECTORY,
                   help=f"Florida GOP directory xlsx (default: {DEFAULT_DIRECTORY})")
    p.add_argument("--no-feedback", action="store_true",
                   help="Skip the feedback loop (don't update County_Master columns).")
    p.add_argument("--no-audit", action="store_true",
                   help="Skip Directory_Audit refresh.")
    p.add_argument("--no-dashboard", action="store_true",
                   help="Skip Dashboard refresh.")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    try:
        import pandas as pd
    except ImportError as e:
        print(f"ERROR: missing dependency ({e}). run: pip install -r requirements.txt", file=sys.stderr)
        return 2

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import _master_io as mio  # type: ignore

    master = Path(args.master)
    if not master.exists():
        print(f"ERROR: master workbook not found: {master}", file=sys.stderr)
        print(f"       copy the boss's florida_sign_operations_master_*.xlsx to {master}",
              file=sys.stderr)
        return 1

    print(f"Refreshing {master} ...")

    if not args.no_feedback:
        print("\n— Feedback loop: operational sheets → County_Master")
        rows = run_feedback_loop(master, mio, pd)
        print(f"  Updated {rows} rows in County_Master.")

    if not args.no_audit:
        print("\n— Directory_Audit")
        directory_path = Path(args.directory)
        if directory_path.exists():
            n = build_directory_audit(master, directory_path, mio, pd)
            print(f"  Wrote {n} audit rows.")
        else:
            print(f"  WARN: {directory_path} missing; skipping.")

    if not args.no_dashboard:
        print("\n— Dashboard")
        build_dashboard(master, mio, pd)
        print("  done.")

    print(f"\nMaster refreshed: {master}")
    return 0


# ──────────────────────────────────────────────────────────────────────────────
# Feedback loop: operational sheets → County_Master columns
# ──────────────────────────────────────────────────────────────────────────────

def run_feedback_loop(master: Path, mio, pd) -> int:
    """Update County_Master.Yard Signs Delivered, Yard Signs Remaining, 4x8 Confirmed."""
    cm = mio.read_sheet_safe(master, "County_Master")
    ysd = mio.read_sheet_safe(master, "Yard_Sign_Deliveries")
    lsl = mio.read_sheet_safe(master, "Large_Sign_Locations")

    if cm.empty or "County" not in cm.columns:
        print("  WARN: County_Master sheet empty or missing County column.")
        return 0

    cm_counties = cm["County"].astype(str).str.strip().tolist()

    # Aggregate yard sign deliveries.
    yard_by_county: dict[str, int] = {}
    if not ysd.empty and "County" in ysd.columns and "Qty Delivered" in ysd.columns:
        agg = ysd.groupby(ysd["County"].astype(str).str.strip())["Qty Delivered"].sum(min_count=1)
        yard_by_county = {k: int(v or 0) for k, v in agg.dropna().items()}
        print(f"  Aggregated Yard_Sign_Deliveries: {len(yard_by_county)} counties with deliveries.")
    else:
        print("  Yard_Sign_Deliveries empty — Yard Signs Delivered will be 0 everywhere.")

    # Aggregate 4x8 installs.
    fourx8_by_county: dict[str, int] = {}
    if not lsl.empty and "County" in lsl.columns and "Install Status" in lsl.columns:
        installed = lsl[lsl["Install Status"].astype(str).str.lower().str.strip() == "installed"]
        agg = installed.groupby(installed["County"].astype(str).str.strip()).size()
        fourx8_by_county = {k: int(v) for k, v in agg.items()}
        print(f"  Aggregated Large_Sign_Locations (Installed): {len(fourx8_by_county)} counties.")
    else:
        print("  Large_Sign_Locations empty — 4x8 Confirmed will be 0 everywhere.")

    # Build per-county updates.
    updates: dict[str, dict[str, object]] = {}
    for _, row in cm.iterrows():
        county = str(row["County"]).strip()
        suggested_yard = row.get("Suggested Yard Signs", 0) or 0
        delivered_yard = yard_by_county.get(county, 0)
        confirmed_4x8 = fourx8_by_county.get(county, 0)
        try:
            remaining = max(int(suggested_yard) - delivered_yard, 0)
        except (TypeError, ValueError):
            remaining = None
        updates[county] = {
            "Yard Signs Delivered": delivered_yard,
            "Yard Signs Remaining": remaining,
            "4x8 Confirmed": confirmed_4x8,
        }
    return mio.update_cells(master, "County_Master", "County", updates)


# ──────────────────────────────────────────────────────────────────────────────
# Directory_Audit: cross-walk florida_gop_directory.xlsx against County_Master
# ──────────────────────────────────────────────────────────────────────────────

def build_directory_audit(master: Path, directory_path: Path, mio, pd) -> int:
    """Build a per-county audit comparing the Directory to County_Master."""
    md = pd.read_excel(directory_path, sheet_name="Master Directory")
    cm = mio.read_sheet_safe(master, "County_Master")

    md["County"] = md["County"].astype(str).str.strip()
    cm["County"] = cm["County"].astype(str).str.strip()

    audit_rows = []
    cm_chairs = dict(zip(cm["County"], cm.get("Chair", "").astype(str).str.strip()))

    for _, cm_row in cm.iterrows():
        county = cm_row["County"]
        cm_chair = str(cm_row.get("Chair") or "").strip()
        sub = md[md["County"] == county]

        # Total directory rows for this county
        n_total = len(sub)
        n_email = sub["Email"].notna().sum() if "Email" in sub.columns else 0
        n_phone = sub["Phone"].notna().sum() if "Phone" in sub.columns else 0
        n_addr  = sub["Address"].notna().sum() if "Address" in sub.columns else 0

        # Chair cross-walk
        chair_in_directory = ""
        if cm_chair and not sub.empty and "Name" in sub.columns:
            names = sub["Name"].astype(str).str.strip()
            if (names == cm_chair).any():
                chair_in_directory = "exact match"
            else:
                # Loose match: surname appears in any directory row for this county
                surname = cm_chair.split()[-1] if cm_chair.split() else ""
                if surname and names.str.contains(surname, regex=False).any():
                    chair_in_directory = f"surname match ({surname})"
                else:
                    chair_in_directory = "MISSING"
        elif not cm_chair:
            chair_in_directory = "(no chair listed in County_Master)"
        else:
            chair_in_directory = "no directory entries for county"

        # Category breakdown
        cats_present = sub["Category"].value_counts().to_dict() if "Category" in sub.columns else {}
        rec_count = cats_present.get("REC Officer", 0)
        club_count = cats_present.get("Club", 0)
        frw_count = cats_present.get("FRW", 0)

        # Recommended action
        gaps = []
        if chair_in_directory == "MISSING":
            gaps.append("verify chair in directory")
        if n_email == 0:
            gaps.append("get email")
        elif n_email / max(n_total, 1) < 0.3:
            gaps.append("low email coverage")
        if n_phone == 0:
            gaps.append("get phone")
        action = "; ".join(gaps) if gaps else "OK"

        audit_rows.append({
            "County": county,
            "Tier": cm_row.get("Strategic Tier", ""),
            "CM_Chair": cm_chair,
            "Chair_in_Directory": chair_in_directory,
            "Directory_Records": n_total,
            "REC_Officers": rec_count,
            "Clubs": club_count,
            "FRW": frw_count,
            "Email_Coverage": f"{n_email}/{n_total}" if n_total else "0/0",
            "Phone_Coverage": f"{n_phone}/{n_total}" if n_total else "0/0",
            "Address_Coverage": f"{n_addr}/{n_total}" if n_total else "0/0",
            "Recommended_Action": action,
        })

    audit_df = pd.DataFrame(audit_rows).sort_values(
        ["Tier", "County"], na_position="last"
    ).reset_index(drop=True)

    mio.replace_sheet(
        master, "Directory_Audit", audit_df,
        color_col="Recommended_Action",
        color_map={"OK": "D4EDDA"},
    )
    return len(audit_df)


# ──────────────────────────────────────────────────────────────────────────────
# Dashboard
# ──────────────────────────────────────────────────────────────────────────────

def build_dashboard(master: Path, mio, pd):
    from openpyxl.styles import Font, PatternFill, Alignment

    cm = mio.read_sheet_safe(master, "County_Master")
    ca = mio.read_sheet_safe(master, "County_Audit")
    pcl = mio.read_sheet_safe(master, "Priority_Call_List")
    olog = mio.read_sheet_safe(master, "Outreach_Log")
    ysd = mio.read_sheet_safe(master, "Yard_Sign_Deliveries")
    lsl = mio.read_sheet_safe(master, "Large_Sign_Locations")
    slc = mio.read_sheet_safe(master, "Sign_Location_Candidates")
    sdp = mio.read_sheet_safe(master, "Sign_Deployment_Plan")
    ysa = mio.read_sheet_safe(master, "Yard_Sign_Allocation")

    # Sign totals
    suggested_yard = int(cm["Suggested Yard Signs"].fillna(0).sum()) if "Suggested Yard Signs" in cm.columns else 0
    delivered_yard = int(cm["Yard Signs Delivered"].fillna(0).sum()) if "Yard Signs Delivered" in cm.columns else 0
    suggested_4x8 = int(cm["Suggested 4x8 Goal"].fillna(0).sum()) if "Suggested 4x8 Goal" in cm.columns else 0
    confirmed_4x8 = int(cm["4x8 Confirmed"].fillna(0).sum()) if "4x8 Confirmed" in cm.columns else 0

    # Tier counts
    tier_counts = cm["Strategic Tier"].value_counts().to_dict() if "Strategic Tier" in cm.columns else {}

    # Completeness
    counties_complete = int((ca["Completeness Score"] >= 80).sum()) if "Completeness Score" in ca.columns else 0
    counties_under = int((ca["Completeness Score"] < 80).sum()) if "Completeness Score" in ca.columns else 0
    counties_total = len(ca) if not ca.empty else len(cm)

    # Filter empty rows (boss's operational sheets are pre-formatted with 200 empty rows).
    def non_empty_rows(df, key_col):
        if df.empty or key_col not in df.columns:
            return df
        return df[df[key_col].notna() & (df[key_col].astype(str).str.strip() != "")]

    ysd = non_empty_rows(ysd, "County")
    lsl = non_empty_rows(lsl, "County")
    olog = non_empty_rows(olog, "County")

    # Outreach
    outreach_rows = len(olog)
    open_followups = 0
    if not olog.empty and "Result Status" in olog.columns:
        open_followups = int((olog["Result Status"].astype(str).str.lower().isin(
            ["open", "pending", "follow-up", "followup"])).sum())

    # Top 5 priority counties (Priority_Call_List)
    top_priority = []
    if not pcl.empty and "Priority Bucket" in pcl.columns:
        pcl_sorted = pcl.sort_values("Priority Bucket")
        top_priority = pcl_sorted.head(5)[["County", "Priority Bucket", "Strategic Tier",
                                            "Completeness Score", "Recommended Ask"]].values.tolist()

    # Build the sheet
    wb = mio._load_and_freeze(master)
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard", 0)

    title = Font(bold=True, size=14)
    h2 = Font(bold=True, size=11)
    label_align = Alignment(horizontal="left", vertical="center")

    ws["A1"] = "Florida Sign Operations — Master Dashboard"
    ws["A1"].font = title
    ws["A2"] = f"Last refreshed: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells("A1:E1")

    def section(row, label):
        ws.cell(row=row, column=1, value=label).font = h2
        return row + 1

    def kv(row, k, v):
        ws.cell(row=row, column=1, value=k).alignment = label_align
        ws.cell(row=row, column=2, value=v)
        return row + 1

    r = 4
    r = section(r, "Sign Deployment (statewide)")
    r = kv(r, "Yard Signs — suggested", suggested_yard)
    r = kv(r, "Yard Signs — delivered", f"{delivered_yard} ({_pct(delivered_yard, suggested_yard)})")
    r = kv(r, "Yard Signs — remaining", suggested_yard - delivered_yard)
    r = kv(r, "4x8 — goal", suggested_4x8)
    r = kv(r, "4x8 — confirmed", f"{confirmed_4x8} ({_pct(confirmed_4x8, suggested_4x8)})")
    r += 1

    r = section(r, "County Coverage")
    r = kv(r, "Counties total", len(cm))
    r = kv(r, "Counties Tier A", tier_counts.get("A", 0))
    r = kv(r, "Counties Tier B", tier_counts.get("B", 0))
    r = kv(r, "Counties Tier C", tier_counts.get("C", 0))
    r = kv(r, "Counties Tier D", tier_counts.get("D", 0))
    r = kv(r, "Counties completeness ≥ 80", f"{counties_complete} / {counties_total}")
    r = kv(r, "Counties completeness < 80", f"{counties_under} / {counties_total}")
    r += 1

    r = section(r, "Outreach Pipeline")
    r = kv(r, "Outreach_Log entries", outreach_rows)
    r = kv(r, "Open / pending follow-ups", open_followups)
    r += 1

    r = section(r, "Operational Activity")
    r = kv(r, "Yard_Sign_Deliveries rows", len(ysd))
    r = kv(r, "Large_Sign_Locations rows", len(lsl))
    if not lsl.empty and "Install Status" in lsl.columns:
        installed = int((lsl["Install Status"].astype(str).str.lower().str.strip() == "installed").sum())
        r = kv(r, "  of which Installed", installed)
    r += 1

    r = section(r, "4×8 Sign Location Candidates (raw pool)")
    r = kv(r, "Total candidates (Sign_Location_Candidates)", len(slc))
    if not slc.empty and "Type" in slc.columns:
        type_counts = slc["Type"].value_counts().to_dict()
        for tname in ("intersection", "commercial", "agri_civic"):
            r = kv(r, f"  {tname}", type_counts.get(tname, 0))
    if not slc.empty and "Phone" in slc.columns:
        ph = (slc["Phone"].fillna("").astype(str).str.len() > 0).sum()
        r = kv(r, "  with verified phone", f"{int(ph)} / {len(slc)}")
    r += 1

    r = section(r, "4×8 Deployment Plan (Sign_Deployment_Plan)")
    if not sdp.empty:
        n_primary = int((sdp["Plan"] == "primary").sum())
        n_replace = int((sdp["Plan"] == "replacement").sum())
        r = kv(r, "Primary deployment", n_primary)
        r = kv(r, "Replacement bench", n_replace)
        prim = sdp[sdp["Plan"] == "primary"]
        if len(prim):
            ph = int((prim["Phone"].fillna("").astype(str).str.len() > 0).sum())
            em = int((prim["Email"].fillna("").astype(str).str.len() > 0).sum())
            r = kv(r, "Primary with phone",
                   f"{ph} / {len(prim)} ({_pct(ph, len(prim))})")
            r = kv(r, "Primary with email",
                   f"{em} / {len(prim)} ({_pct(em, len(prim))})")
            r = kv(r, "Primary by Tier (count)",
                   ", ".join(f"{t}={int(n)}" for t, n in
                             prim["Tier"].value_counts().sort_index().items()))
    else:
        r = kv(r, "(empty — run select_sign_deployment.py)", "")
    r += 1

    r = section(r, "Yard Sign Allocation (Yard_Sign_Allocation)")
    if not ysa.empty:
        boss_total = int(ysa["Suggested_Yard_Signs_Boss"].fillna(0).sum())
        plan_total = int(ysa["Plan_4000"].fillna(0).sum())
        r = kv(r, "Boss plan total (7,485)", boss_total)
        r = kv(r, "Plan_4000 total (THE NUMBER — physical inventory)", plan_total)
        for tier in ["A", "B", "C", "D"]:
            sub = ysa[ysa["Strategic_Tier"] == tier]
            if len(sub):
                avg = int(sub['Plan_4000'].mean()) if len(sub) else 0
                r = kv(r, f"  Tier {tier}: counties / Plan_4000 / avg per",
                       f"{len(sub)} counties, {int(sub['Plan_4000'].sum())} signs, avg {avg}/REC")
        complete = int((ysa["POC_Status_Yard"] == "chair_complete").sum())
        r = kv(r, "  Counties with full chair POC", f"{complete} / {len(ysa)}")
    else:
        r = kv(r, "(empty — run allocate_yard_signs.py)", "")
    r += 1

    # ── Outreach Pipeline (4x8 call queue) ────────────────────────────────────
    r = section(r, "📞 Outreach Pipeline (4x8 Call Queue)")
    if not olog.empty and "Result Status" in olog.columns:
        log4x8 = olog[olog.get("Contact Type", pd.Series("", index=olog.index)) == "4x8 Site Owner"]
        if len(log4x8):
            status_counts = log4x8["Result Status"].fillna("(blank)").value_counts()
            r = kv(r, "Total 4x8 outreach rows", len(log4x8))
            for status in ["To Call", "Voicemail", "No Answer", "Reached - Pending",
                           "Approved", "Declined", "Wrong Number", "Do Not Contact", "Closed"]:
                n = int(status_counts.get(status, 0))
                if n > 0:
                    r = kv(r, f"  {status}", n)
            approved = int(status_counts.get("Approved", 0))
            declined = int(status_counts.get("Declined", 0))
            called = int((log4x8["Result Status"].astype(str) != "To Call").sum())
            r = kv(r, "Calls completed", f"{called} / {len(log4x8)} ({_pct(called, len(log4x8))})")
            r = kv(r, "Approval rate (of completed)",
                   f"{approved}/{called - 0 if called else 0} ({_pct(approved, max(approved+declined, 1))})")
        else:
            r = kv(r, "(no 4x8 outreach rows yet)", "")
    else:
        r = kv(r, "(Outreach_Log empty — run seed_operational_sheets.py)", "")
    r += 1

    # ── Yard Delivery Pipeline ────────────────────────────────────────────────
    r = section(r, "🚚 Yard Delivery Pipeline (REC Coordination)")
    if not olog.empty and "Result Status" in olog.columns:
        rec_log = olog[olog.get("Contact Type", pd.Series("", index=olog.index)) == "REC Chair / Delivery"]
        if len(rec_log):
            status_counts = rec_log["Result Status"].fillna("(blank)").value_counts()
            r = kv(r, "Total REC outreach rows", len(rec_log))
            for status in ["To Call", "Voicemail", "Reached - Pending", "Approved", "Declined"]:
                n = int(status_counts.get(status, 0))
                if n > 0:
                    r = kv(r, f"  {status}", n)
            confirmed = int(status_counts.get("Approved", 0))
            r = kv(r, "RECs confirmed", f"{confirmed} / {len(rec_log)}")

    # Delivery sheet rollup (if it has a Status column)
    if not ysd.empty and "Status" in ysd.columns:
        ds_counts = ysd["Status"].fillna("(blank)").value_counts()
        for status in ["Scheduled", "Confirmed", "In Transit", "Delivered", "Issue", "Cancelled"]:
            n = int(ds_counts.get(status, 0))
            if n > 0:
                r = kv(r, f"  Delivery: {status}", n)
        delivered_qty = 0
        if "Qty Delivered" in ysd.columns:
            delivered_qty = int(pd.to_numeric(ysd["Qty Delivered"], errors="coerce").fillna(0).sum())
        plan_qty = int(ysa["Plan_4000"].fillna(0).sum()) if not ysa.empty else 4000
        r = kv(r, "Yard signs delivered (sum of Qty Delivered)",
               f"{delivered_qty} / {plan_qty} ({_pct(delivered_qty, plan_qty)})")
    r += 1

    r = section(r, "Top Priority Counties (from Priority_Call_List)")
    if top_priority:
        ws.cell(row=r, column=1, value="County"); ws.cell(row=r, column=2, value="Bucket")
        ws.cell(row=r, column=3, value="Tier"); ws.cell(row=r, column=4, value="Score")
        ws.cell(row=r, column=5, value="Recommended Ask")
        for c in range(1, 6):
            ws.cell(row=r, column=c).font = h2
        r += 1
        for row in top_priority:
            for ci, v in enumerate(row, start=1):
                ws.cell(row=r, column=ci, value=v if not (isinstance(v, float) and v != v) else None)
            r += 1
    else:
        ws.cell(row=r, column=1, value="(Priority_Call_List empty)"); r += 1

    # Column widths
    widths = [38, 22, 12, 12, 60]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + col_idx - 1)].width = w

    # Reorder: Executive_Summary first (governor-facing), then Dashboard, Directory_Audit.
    leading = [n for n in ("Executive_Summary", "Dashboard", "Directory_Audit") if n in wb.sheetnames]
    extras = [s for s in wb.sheetnames if s not in leading]
    wb._sheets = [wb[name] for name in leading + extras]

    wb.save(master)


def _pct(n, d) -> str:
    if not d:
        return "0%"
    return f"{(n / d) * 100:.0f}%"


if __name__ == "__main__":
    sys.exit(main())
