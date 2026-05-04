#!/usr/bin/env python3
"""
Phase 0 — Master Workbook Orchestrator

Initializes outputs/campaign_ops_master.xlsx, runs whichever phases have
inputs available, and rebuilds the Inventory + Dashboard sheets.

Usage:
    python scripts/00_build_master.py
    python scripts/00_build_master.py --refresh-only   # rebuild Dashboard + Inventory only
    python scripts/00_build_master.py --with-map       # also refresh sign_map.html
"""
from __future__ import annotations

import argparse
import importlib.util
import sys
from datetime import datetime
from pathlib import Path

# Local helper -- imported lazily inside main() to keep --help fast.

DEFAULT_MASTER = "outputs/campaign_ops_master.xlsx"
DEFAULT_CONTACTS = "data/raw/contacts.xlsx"
DEFAULT_T22 = "data/raw/turnout_2022.csv"
DEFAULT_T24 = "data/raw/turnout_2024.csv"


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Phase 0 — Build / refresh the master campaign workbook.")
    p.add_argument("--master", default=DEFAULT_MASTER, help=f"Master xlsx path (default: {DEFAULT_MASTER})")
    p.add_argument("--contacts-input", default=DEFAULT_CONTACTS)
    p.add_argument("--turnout-2022", default=DEFAULT_T22)
    p.add_argument("--turnout-2024", default=DEFAULT_T24)
    p.add_argument("--refresh-only", action="store_true",
                   help="Skip phase scripts; only rebuild Inventory + Dashboard.")
    p.add_argument("--with-map", action="store_true",
                   help="Also re-run Phase 3 (GIS map). Slow.")
    p.add_argument("--shapefile", default=None,
                   help="Required if --with-map is set; path to county precinct .shp.")
    return p.parse_args()


def _import_phase(script_name: str):
    """Import a numbered phase script as a module (its filename starts with a digit)."""
    here = Path(__file__).resolve().parent
    target = here / script_name
    spec = importlib.util.spec_from_file_location(script_name.replace(".py", ""), target)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def main() -> int:
    args = parse_args()

    try:
        import _master_io as mio
    except ImportError:
        # When run as `python scripts/00_build_master.py`, scripts/ may not be on sys.path.
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        import _master_io as mio  # type: ignore

    try:
        import pandas as pd
    except ImportError as e:
        print(f"ERROR: missing dependency ({e}). run: pip install -r requirements.txt", file=sys.stderr)
        return 2

    master = Path(args.master)
    print(f"Initializing master workbook at {master} ...")
    mio.init_master(master)

    if not args.refresh_only:
        # Phase 1
        if Path(args.contacts_input).exists():
            print("\n— Phase 1 — Contact audit")
            phase1 = _import_phase("01_audit_contacts.py")
            sys.argv = ["01_audit_contacts.py",
                        "--input", args.contacts_input,
                        "--master", str(master),
                        "--auto-confirm"]
            try:
                phase1.main()
            except SystemExit as e:
                if e.code not in (0, None):
                    print(f"  Phase 1 exited with code {e.code}; continuing.")
        else:
            print(f"\n— Phase 1 skipped: {args.contacts_input} not found.")

        # Phase 2
        if Path(args.turnout_2022).exists() and Path(args.turnout_2024).exists():
            print("\n— Phase 2 — Turnout analysis")
            phase2 = _import_phase("02_turnout_analysis.py")
            sys.argv = ["02_turnout_analysis.py",
                        "--turnout-2022", args.turnout_2022,
                        "--turnout-2024", args.turnout_2024,
                        "--master", str(master)]
            try:
                phase2.main()
            except SystemExit as e:
                if e.code not in (0, None):
                    print(f"  Phase 2 exited with code {e.code}; continuing.")
        else:
            print("\n— Phase 2 skipped: turnout CSVs not found.")

        # Phase 4 -- only if both Phase 1 and Phase 2 produced data.
        contacts_df = mio.read_sheet_safe(master, "Contacts")
        precincts_df = mio.read_sheet_safe(master, "Precincts")
        if not contacts_df.empty and not precincts_df.empty:
            print("\n— Phase 4 — Sign allocation")
            phase4 = _import_phase("04_sign_allocation.py")
            sys.argv = ["04_sign_allocation.py", "--master", str(master)]
            try:
                phase4.main()
            except SystemExit as e:
                if e.code not in (0, None):
                    print(f"  Phase 4 exited with code {e.code}; continuing.")
        else:
            print("\n— Phase 4 skipped: Contacts or Precincts sheet empty.")

    # ── Inventory rebuild ────────────────────────────────────────────────────
    print("\n— Rebuilding Inventory")
    inv = build_inventory(master, mio, pd)
    if not inv.empty:
        mio.replace_sheet(master, "Inventory", inv)
        print(f"  {len(inv)} REC rows.")
    else:
        print("  no allocation data; Inventory left empty.")

    # ── Dashboard rebuild ────────────────────────────────────────────────────
    print("\n— Rebuilding Dashboard")
    build_dashboard(master, mio, pd)
    print("  done.")

    # ── Optional: Phase 3 map ────────────────────────────────────────────────
    if args.with_map:
        if not args.shapefile:
            print("WARN: --with-map requires --shapefile; skipping map regeneration.")
        else:
            print("\n— Phase 3 — GIS map")
            phase3 = _import_phase("03_gis_map.py")
            sys.argv = ["03_gis_map.py",
                        "--shapefile", args.shapefile,
                        "--master", str(master)]
            try:
                phase3.main()
            except SystemExit as e:
                if e.code not in (0, None):
                    print(f"  Phase 3 exited with code {e.code}.")

    print(f"\nMaster workbook ready: {master}")
    return 0


def build_inventory(master: Path, mio, pd):
    """Join Allocation × Sign_Tracker into a per-REC inventory table."""
    alloc = mio.read_sheet_safe(master, "Allocation")
    tracker = mio.read_sheet_safe(master, "Sign_Tracker")
    if alloc.empty:
        return pd.DataFrame()

    # Tracker counts by REC + Type + Status.
    if not tracker.empty and "REC" in tracker.columns:
        t = tracker.copy()
        t["Type"] = t.get("Type", "").astype(str).str.lower()
        t["Status"] = t.get("Status", "").astype(str).str.lower()
        delivered_4x8 = t[(t["Type"] == "4x8") & (t["Status"].isin(["delivered", "installed"]))].groupby("REC").size()
        installed_4x8 = t[(t["Type"] == "4x8") & (t["Status"] == "installed")].groupby("REC").size()
        delivered_yard = t[(t["Type"] == "yard") & (t["Status"].isin(["delivered", "installed"]))].groupby("REC").size()
        installed_yard = t[(t["Type"] == "yard") & (t["Status"] == "installed")].groupby("REC").size()
    else:
        delivered_4x8 = installed_4x8 = delivered_yard = installed_yard = pd.Series(dtype=int)

    rows = []
    for _, a in alloc.iterrows():
        rec = a.get("REC_Name", "")
        rows.append({
            "REC": rec,
            "County": a.get("County", ""),
            "Quality": a.get("Quality", ""),
            "4x8_Allocated": int(a.get("4x8_qty", 0) or 0),
            "4x8_Delivered": int(delivered_4x8.get(rec, 0)),
            "4x8_Installed": int(installed_4x8.get(rec, 0)),
            "Yard_Allocated": int(a.get("yard_qty", 0) or 0),
            "Yard_Delivered": int(delivered_yard.get(rec, 0)),
            "Yard_Installed": int(installed_yard.get(rec, 0)),
        })
    inv = pd.DataFrame(rows)
    inv["4x8_Remaining"] = inv["4x8_Allocated"] - inv["4x8_Delivered"]
    inv["Yard_Remaining"] = inv["Yard_Allocated"] - inv["Yard_Delivered"]

    def status(r):
        if r["Quality"] == "Broken":
            return "Blocked"
        total_alloc = r["4x8_Allocated"] + r["Yard_Allocated"]
        total_deliv = r["4x8_Delivered"] + r["Yard_Delivered"]
        if total_alloc > 0 and total_deliv < 0.5 * total_alloc:
            return "Behind"
        return "OK"
    inv["Status"] = inv.apply(status, axis=1)

    return inv[[
        "REC", "County", "Quality",
        "4x8_Allocated", "4x8_Delivered", "4x8_Installed", "4x8_Remaining",
        "Yard_Allocated", "Yard_Delivered", "Yard_Installed", "Yard_Remaining",
        "Status",
    ]]


def build_dashboard(master: Path, mio, pd):
    """Rebuild the Dashboard sheet with rolled-up KPIs."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    contacts = mio.read_sheet_safe(master, "Contacts")
    precincts = mio.read_sheet_safe(master, "Precincts")
    alloc = mio.read_sheet_safe(master, "Allocation")
    tracker = mio.read_sheet_safe(master, "Sign_Tracker")
    delivery = mio.read_sheet_safe(master, "Delivery_Log")
    followup = mio.read_sheet_safe(master, "Followup_Log")
    inv = mio.read_sheet_safe(master, "Inventory")

    def s(df, col): return df[col].sum() if not df.empty and col in df.columns else 0

    total_4x8_alloc = int(s(alloc, "4x8_qty"))
    total_yard_alloc = int(s(alloc, "yard_qty"))
    total_4x8_deliv = int(s(inv, "4x8_Delivered"))
    total_4x8_inst = int(s(inv, "4x8_Installed"))
    total_yard_deliv = int(s(inv, "Yard_Delivered"))
    total_yard_inst = int(s(inv, "Yard_Installed"))

    rec_total = len(alloc)
    rec_confirmed = int((alloc["Confirmed_Y_N"] == "Y").sum()) if "Confirmed_Y_N" in alloc.columns else 0
    rec_blocked = int((alloc["Quality"] == "Broken").sum()) if "Quality" in alloc.columns else 0
    rec_behind = int((inv["Status"] == "Behind").sum()) if "Status" in inv.columns else 0

    # Tier coverage
    def tier_coverage(tier_n: int) -> str:
        if precincts.empty or "tier" not in precincts.columns:
            return "n/a"
        tier_precincts = precincts[precincts["tier"] == tier_n]
        if tier_precincts.empty:
            return "n/a"
        if tracker.empty or "Precinct" not in tracker.columns:
            return f"0 / {len(tier_precincts)} (0%)"
        installed = tracker[tracker.get("Status", "").astype(str).str.lower() == "installed"]
        covered = installed["Precinct"].astype(str).isin(tier_precincts["precinct_id"].astype(str)).sum()
        pct = covered / len(tier_precincts) * 100 if len(tier_precincts) else 0
        return f"{covered} / {len(tier_precincts)} ({pct:.0f}%)"

    # Follow-up queue
    overdue = 0
    if not followup.empty and "Due_Date" in followup.columns:
        try:
            due = pd.to_datetime(followup["Due_Date"], errors="coerce")
            today = pd.Timestamp(datetime.now().date())
            overdue = int(((due <= today) & (followup.get("Outcome", "").astype(str).str.strip() == "")).sum())
        except Exception:
            overdue = 0

    # Recent deliveries
    recent_rows = []
    if not delivery.empty and "Date" in delivery.columns:
        recent = delivery.copy()
        recent["__sort__"] = pd.to_datetime(recent["Date"], errors="coerce")
        recent = recent.sort_values("__sort__", ascending=False).head(5)
        for _, r in recent.iterrows():
            recent_rows.append([str(r.get("Date", "")), str(r.get("REC", "")),
                                int(r.get("Quantity_4x8", 0) or 0),
                                int(r.get("Quantity_Yard", 0) or 0),
                                str(r.get("Notes", ""))])

    # Open & build the Dashboard sheet
    wb = load_workbook(master)
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard", 0)  # leftmost

    title = Font(bold=True, size=14)
    h2 = Font(bold=True, size=11)
    pct_fill = PatternFill("solid", fgColor="EAF3FB")

    ws["A1"] = "Campaign Ops — Master Dashboard"
    ws["A1"].font = title
    ws["A2"] = f"Last refreshed: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells("A1:E1")

    def write_row(row_idx, label, value, h2_label=False):
        ws.cell(row=row_idx, column=1, value=label).font = h2 if h2_label else Font()
        ws.cell(row=row_idx, column=2, value=value)

    r = 4
    ws.cell(row=r, column=1, value="Sign Deployment").font = h2; r += 1
    write_row(r, "4x8 allocated",  total_4x8_alloc); r += 1
    write_row(r, "4x8 delivered",  f"{total_4x8_deliv} ({_pct(total_4x8_deliv, total_4x8_alloc)})"); r += 1
    write_row(r, "4x8 installed",  f"{total_4x8_inst} ({_pct(total_4x8_inst, total_4x8_alloc)})"); r += 1
    write_row(r, "Yard allocated", total_yard_alloc); r += 1
    write_row(r, "Yard delivered", f"{total_yard_deliv} ({_pct(total_yard_deliv, total_yard_alloc)})"); r += 1
    write_row(r, "Yard installed", f"{total_yard_inst} ({_pct(total_yard_inst, total_yard_alloc)})"); r += 2

    ws.cell(row=r, column=1, value="REC Status").font = h2; r += 1
    write_row(r, "Total RECs", rec_total); r += 1
    write_row(r, "Confirmed",  rec_confirmed); r += 1
    write_row(r, "Blocked",    rec_blocked); r += 1
    write_row(r, "Behind",     rec_behind); r += 2

    ws.cell(row=r, column=1, value="Tier Coverage").font = h2; r += 1
    write_row(r, "Tier 1 precincts with installed signs", tier_coverage(1)); r += 1
    write_row(r, "Tier 2 precincts with installed signs", tier_coverage(2)); r += 2

    ws.cell(row=r, column=1, value="Follow-up Queue").font = h2; r += 1
    write_row(r, "Overdue follow-ups", overdue); r += 2

    ws.cell(row=r, column=1, value="Recent Deliveries (last 5)").font = h2; r += 1
    ws.cell(row=r, column=1, value="Date"); ws.cell(row=r, column=2, value="REC")
    ws.cell(row=r, column=3, value="4x8"); ws.cell(row=r, column=4, value="Yard")
    ws.cell(row=r, column=5, value="Notes")
    for c in range(1, 6):
        ws.cell(row=r, column=c).font = h2
    r += 1
    if recent_rows:
        for row in recent_rows:
            for ci, v in enumerate(row, start=1):
                ws.cell(row=r, column=ci, value=v)
            r += 1
    else:
        ws.cell(row=r, column=1, value="(none yet)").alignment = Alignment(italic=True) if False else Alignment()

    # Column widths
    for col_idx, w in enumerate([42, 22, 12, 12, 40], start=1):
        ws.column_dimensions[chr(ord("A") + col_idx - 1)].width = w

    # Reorder: Dashboard first
    desired = ["Dashboard"] + [s for s in ["Contacts", "Flagged_Issues", "Precincts",
                                           "Allocation", "Inventory", "Sign_Tracker",
                                           "Delivery_Log", "Followup_Log"] if s in wb.sheetnames]
    extras = [s for s in wb.sheetnames if s not in desired]
    wb._sheets = [wb[name] for name in desired + extras]

    wb.save(master)


def _pct(n, d) -> str:
    if not d:
        return "0%"
    return f"{(n / d) * 100:.0f}%"


if __name__ == "__main__":
    sys.exit(main())
