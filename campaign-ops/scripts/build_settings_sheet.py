#!/usr/bin/env python3
"""
Build / refresh the Settings sheet in the master workbook.

Settings is the single source of truth for top-line totals. The user edits
input cells in Excel; running `regenerate_plans.py` cascades changes downstream.

Layout:
  Inputs (B2:B9):   editable cells the user changes
  Validation (B12:B16): live SUM formulas; show MISMATCH if regenerate is stale

Defined names registered:
  TOTAL_4X8_PRIMARY, TOTAL_YARD_SIGNS, REPLACEMENT_RATIO, MIN_PER_COUNTY_4X8,
  TIER_A_FLOOR, TIER_B_FLOOR, TIER_C_FLOOR, TIER_D_FLOOR

Usage:
    python scripts/build_settings_sheet.py
    python scripts/build_settings_sheet.py --total-4x8 1500 --total-yard 2000
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

DEFAULT_MASTER = "outputs/campaign_ops_master.xlsx"

DEFAULTS = {
    "TOTAL_4X8_PRIMARY":   2000,
    "TOTAL_YARD_SIGNS":    4000,
    "REPLACEMENT_RATIO":   0.25,
    "MIN_PER_COUNTY_4X8":  4,
    "TIER_A_FLOOR":        100,
    "TIER_B_FLOOR":        60,
    "TIER_C_FLOOR":        40,
    "TIER_D_FLOOR":        25,
}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build/refresh the Settings sheet.")
    p.add_argument("--master", default=DEFAULT_MASTER)
    p.add_argument("--total-4x8", type=int, default=None)
    p.add_argument("--total-yard", type=int, default=None)
    p.add_argument("--replacement-ratio", type=float, default=None)
    p.add_argument("--preserve-existing", action="store_true",
                   help="Read existing Settings values from sheet; don't reset to defaults.")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.workbook.defined_name import DefinedName
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        print(f"ERROR: missing dependency ({e}).", file=sys.stderr)
        return 2

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import _master_io as mio  # type: ignore

    master = Path(args.master)
    if not master.exists():
        print(f"ERROR: master not found: {master}", file=sys.stderr)
        return 1

    wb = load_workbook(master)

    # Determine values: existing > CLI args > defaults
    values = dict(DEFAULTS)
    if args.preserve_existing and "Settings" in wb.sheetnames:
        existing = _read_existing_values(wb)
        for k, v in existing.items():
            if v is not None:
                values[k] = v
    if args.total_4x8 is not None:
        values["TOTAL_4X8_PRIMARY"] = args.total_4x8
    if args.total_yard is not None:
        values["TOTAL_YARD_SIGNS"] = args.total_yard
    if args.replacement_ratio is not None:
        values["REPLACEMENT_RATIO"] = args.replacement_ratio

    # Drop existing sheet
    if "Settings" in wb.sheetnames:
        del wb["Settings"]
    # Drop existing defined names that we own
    owned_names = list(DEFAULTS.keys())
    for name in owned_names:
        try:
            if name in wb.defined_names:
                del wb.defined_names[name]
        except Exception:
            pass

    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False

    # ── Style palette ─────────────────────────────────────────────────────────
    title_font = Font(bold=True, size=18, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1F4E79")
    h2 = Font(bold=True, size=13, color="1F4E79")
    label_font = Font(size=11)
    input_font = Font(bold=True, size=12, color="1F4E79")
    input_fill = PatternFill("solid", fgColor="FFF8E1")
    formula_font = Font(size=11, color="555555", italic=True)
    formula_fill = PatternFill("solid", fgColor="EAF3FB")
    note = Font(size=10, color="555555", italic=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # ── Title bar ─────────────────────────────────────────────────────────────
    ws["A1"] = "Settings — Sign Quantities & Distribution Parameters"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30

    # ── Inputs zone ───────────────────────────────────────────────────────────
    ws["A3"] = "INPUTS  (edit these cells, then run: python scripts/regenerate_plans.py)"
    ws["A3"].font = h2
    ws.merge_cells("A3:D3")

    inputs = [
        ("TOTAL_4X8_PRIMARY",   "Total 4×8 signs to deploy",                values["TOTAL_4X8_PRIMARY"],   "Hard inventory cap"),
        ("TOTAL_YARD_SIGNS",    "Total yard signs to deploy",               values["TOTAL_YARD_SIGNS"],   "Hard inventory cap"),
        ("REPLACEMENT_RATIO",   "Replacement bench ratio",                  values["REPLACEMENT_RATIO"],  "0.25 = 25% of primary count as backup"),
        ("MIN_PER_COUNTY_4X8",  "Minimum 4×8 sites per county",             values["MIN_PER_COUNTY_4X8"], "Smallest counties get at least this many"),
        ("TIER_A_FLOOR",        "Yard sign floor — Tier A (urban metros)",  values["TIER_A_FLOOR"],       ""),
        ("TIER_B_FLOOR",        "Yard sign floor — Tier B",                 values["TIER_B_FLOOR"],       ""),
        ("TIER_C_FLOOR",        "Yard sign floor — Tier C",                 values["TIER_C_FLOOR"],       ""),
        ("TIER_D_FLOOR",        "Yard sign floor — Tier D (rural)",         values["TIER_D_FLOOR"],       ""),
    ]
    r = 4
    # Header row
    ws.cell(row=r, column=1, value="Name").font = Font(bold=True, color="555555")
    ws.cell(row=r, column=2, value="Value").font = Font(bold=True, color="555555")
    ws.cell(row=r, column=3, value="Description").font = Font(bold=True, color="555555")
    ws.cell(row=r, column=4, value="Note").font = Font(bold=True, color="555555")
    r += 1
    input_first_row = r
    for name, desc, val, hint in inputs:
        ws.cell(row=r, column=1, value=name).font = label_font
        cell = ws.cell(row=r, column=2, value=val)
        cell.font = input_font
        cell.fill = input_fill
        cell.alignment = center
        # Register defined name pointing to this cell
        ref = f"Settings!${get_column_letter(2)}${r}"
        try:
            dn = DefinedName(name=name, attr_text=ref)
            wb.defined_names[name] = dn
        except Exception:
            pass
        ws.cell(row=r, column=3, value=desc).font = label_font
        ws.cell(row=r, column=4, value=hint).font = note
        r += 1

    # ── Validation zone ───────────────────────────────────────────────────────
    r += 2
    ws.cell(row=r, column=1, value="VALIDATION  (live formulas — refresh in Excel after regenerate)").font = h2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1

    val_first_row = r
    # Yard plan total
    ws.cell(row=r, column=1, value="Yard_Plan_Sum").font = label_font
    ws.cell(row=r, column=2, value="=SUMIFS(Yard_Sign_Allocation!H2:H100,Yard_Sign_Allocation!A2:A100,\"<>\")").font = formula_font
    ws.cell(row=r, column=2).fill = formula_fill
    ws.cell(row=r, column=3, value="Sum of Plan_4000 across all 67 counties").font = label_font
    r += 1
    ws.cell(row=r, column=1, value="Yard_OK").font = label_font
    ws.cell(row=r, column=2,
            value=f'=IF(B{val_first_row}=TOTAL_YARD_SIGNS,"✓ OK","✗ MISMATCH — run regenerate_plans")').font = formula_font
    ws.cell(row=r, column=2).fill = formula_fill
    ws.cell(row=r, column=3, value="Should equal TOTAL_YARD_SIGNS").font = label_font
    r += 1
    # Primary count
    ws.cell(row=r, column=1, value="Primary_Count").font = label_font
    ws.cell(row=r, column=2, value="=COUNTIF(Sign_Deployment_Plan!A:A,\"primary\")").font = formula_font
    ws.cell(row=r, column=2).fill = formula_fill
    ws.cell(row=r, column=3, value="Count of primary 4×8 sites").font = label_font
    r += 1
    ws.cell(row=r, column=1, value="Primary_OK").font = label_font
    ws.cell(row=r, column=2,
            value=f'=IF(B{r-1}=TOTAL_4X8_PRIMARY,"✓ OK","✗ MISMATCH — run regenerate_plans")').font = formula_font
    ws.cell(row=r, column=2).fill = formula_fill
    ws.cell(row=r, column=3, value="Should equal TOTAL_4X8_PRIMARY").font = label_font
    r += 2

    ws.cell(row=r, column=1, value="Last_Regenerated").font = label_font
    ws.cell(row=r, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")).font = formula_font
    ws.cell(row=r, column=3, value="Updated by regenerate_plans.py").font = note

    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 38

    # Pin Settings just after Executive_Summary
    leading = []
    for n in ("Executive_Summary", "Settings", "Dashboard", "Directory_Audit"):
        if n in wb.sheetnames:
            leading.append(n)
    extras = [s for s in wb.sheetnames if s not in leading]
    wb._sheets = [wb[name] for name in leading + extras]

    wb.save(master)
    print(f"Settings sheet written. Inputs:")
    for k, v in values.items():
        print(f"  {k:25s} {v}")
    return 0


def _read_existing_values(wb) -> dict:
    """Read current values from existing Settings sheet (best-effort)."""
    out = {}
    if "Settings" not in wb.sheetnames:
        return out
    ws = wb["Settings"]
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if name in DEFAULTS:
            v = row[1]
            if v is None:
                continue
            try:
                if isinstance(DEFAULTS[name], int):
                    out[name] = int(v)
                elif isinstance(DEFAULTS[name], float):
                    out[name] = float(v)
                else:
                    out[name] = v
            except (TypeError, ValueError):
                pass
    return out


if __name__ == "__main__":
    sys.exit(main())
