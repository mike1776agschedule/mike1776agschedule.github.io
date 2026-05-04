#!/usr/bin/env python3
"""
Build a simplified action workbook for daily field ops.

Source: outputs/campaign_ops_master.xlsx (24 sheets — full reference workbook)
Output: outputs/campaign_ops_action.xlsx (5 focused sheets)

Sheets:
  1. Start_Here           Cover page with KPIs + how-to-use
  2. 4x8_Sites            2,000 primary 4×8 sites — the call list (filterable, status dropdowns)
  3. Yard_Sign_RECs       67 REC delivery rows with chair contacts + Plan_4000 quantities
  4. Call_Queue           2,067 outreach rows merged from Outreach_Log
  5. Replacement_Bench    480 backup sites for refusal recovery

Usage:
    python scripts/build_action_workbook.py
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

DEFAULT_MASTER = "outputs/campaign_ops_master.xlsx"
DEFAULT_OUTPUT = "outputs/campaign_ops_action.xlsx"


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build simplified action workbook from master.")
    p.add_argument("--master", default=DEFAULT_MASTER)
    p.add_argument("--output", default=DEFAULT_OUTPUT)
    return p.parse_args()


def main() -> int:
    args = parse_args()
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as e:
        print(f"ERROR: missing dependency ({e}).", file=sys.stderr)
        return 2

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import _master_io as mio  # type: ignore

    master = Path(args.master)
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)

    print(f"Reading {master} ...")
    sdp = mio.read_sheet_safe(master, "Sign_Deployment_Plan")
    ysa = mio.read_sheet_safe(master, "Yard_Sign_Allocation")
    olog = mio.read_sheet_safe(master, "Outreach_Log")
    cm = mio.read_sheet_safe(master, "County_Master")

    # Idempotent edit-preservation: read existing action workbook (if any) and
    # build per-Candidate_ID / per-County overlay of field-staff-edited fields.
    sites_overlay: dict[str, dict] = {}
    recs_overlay: dict[str, dict] = {}
    if out.exists():
        try:
            old_sites = pd.read_excel(out, sheet_name="4x8_Sites")
            for _, r in old_sites.iterrows():
                cid = str(r.get("Candidate_ID") or "").strip()
                if not cid:
                    continue
                kept = {}
                for col in ("Phone", "Email", "Owner_Contact", "Notes",
                            "Approval_Status", "Install_Status"):
                    val = r.get(col)
                    if val is None: continue
                    if isinstance(val, float) and val != val: continue
                    sval = str(val).strip()
                    if sval and sval.lower() not in ("nan", "none", "null", "pending"):
                        kept[col] = sval
                if kept:
                    sites_overlay[cid] = kept
            old_recs = pd.read_excel(out, sheet_name="Yard_Sign_RECs")
            for _, r in old_recs.iterrows():
                county = str(r.get("County") or "").strip()
                if not county: continue
                kept = {}
                for col in ("Status", "Delivery_Date", "Notes",
                            "Chair_Phone", "Chair_Email"):
                    val = r.get(col)
                    if val is None: continue
                    if isinstance(val, float) and val != val: continue
                    sval = str(val).strip()
                    if sval and sval.lower() not in ("nan", "none", "null", "scheduled"):
                        kept[col] = sval
                if kept:
                    recs_overlay[county] = kept
            print(f"  Preserving edits: {len(sites_overlay)} site overlays, {len(recs_overlay)} REC overlays")
            # Backup the existing workbook
            bak = out.with_suffix(".bak.xlsx")
            try:
                import shutil
                shutil.copy2(out, bak)
                print(f"  Backup: {bak}")
            except Exception as e:
                print(f"  WARN: backup failed: {e}")
        except Exception as e:
            print(f"  WARN: could not read existing action workbook ({e}); starting fresh")

    if sdp.empty or ysa.empty or olog.empty:
        print("ERROR: master sheets empty. run pipeline first.", file=sys.stderr)
        return 1

    # ── Build new workbook from scratch ─────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)  # drop default empty sheet

    # Style palette
    title_font = Font(bold=True, size=20, color="FFFFFF", name="Calibri")
    title_fill = PatternFill("solid", fgColor="1F4E79")
    h2_font = Font(bold=True, size=14, color="1F4E79")
    h3_font = Font(bold=True, size=11)
    body = Font(size=11)
    muted = Font(size=10, color="555555")
    kpi_value = Font(bold=True, size=22, color="1F4E79")
    kpi_label = Font(size=10, color="555555")
    table_header_font = Font(bold=True, color="FFFFFF")
    table_header_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill = PatternFill("solid", fgColor="F4F8FB")
    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # ── Sheet 1: Start_Here ─────────────────────────────────────────────────
    ws = wb.create_sheet("Start_Here")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Florida Sign Operations — Action Workbook"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 38

    ws["A2"] = f"Generated {datetime.now().strftime('%B %d, %Y')}  |  Field-ops cut from master workbook"
    ws["A2"].font = muted
    ws["A2"].alignment = center
    ws.merge_cells("A2:F2")

    # KPIs
    prim = sdp[sdp["Plan"] == "primary"]
    repl = sdp[sdp["Plan"] == "replacement"]
    n_primary = len(prim)
    n_repl = len(repl)
    yard_total = int(ysa["Plan_4000"].fillna(0).sum())
    chair_complete = int((ysa["POC_Status_Yard"] == "chair_complete").sum())
    n_phone = int((prim["Phone"].fillna("").astype(str).str.len() > 0).sum())

    def kpi(row, col, value, label):
        ws.cell(row=row, column=col, value=value).font = kpi_value
        ws.cell(row=row, column=col).alignment = center
        ws.cell(row=row + 1, column=col, value=label).font = kpi_label
        ws.cell(row=row + 1, column=col).alignment = center

    r = 4
    ws.cell(row=r, column=1, value="HEADLINE NUMBERS").font = h2_font
    r += 1
    kpi(r, 1, f"{n_primary:,}",        "4×8 SITES TO CALL")
    kpi(r, 3, f"{yard_total:,}",       "YARD SIGNS TO DELIVER")
    kpi(r, 5, f"{chair_complete}/67",  "RECs READY TO CONTACT")
    r += 3
    kpi(r, 1, f"{n_repl:,}",           "REPLACEMENT BENCH")
    kpi(r, 3, f"{n_phone:,}",          "SITES WITH VERIFIED PHONE")
    kpi(r, 5, "67",                    "FLORIDA COUNTIES")
    r += 4

    # How to use
    ws.cell(row=r, column=1, value="HOW TO USE THIS WORKBOOK").font = h2_font
    r += 1
    workflow = [
        ("4x8_Sites",
         "Filter by County, sort by Composite_Score (highest = call first). "
         "Set Install_Status as you progress: Pending → Approved → Installed."),
        ("Yard_Sign_RECs",
         "Call each REC chair to schedule yard sign delivery. Update Status: "
         "Scheduled → Confirmed → In Transit → Delivered. Quantity is in Plan_4000."),
        ("Call_Queue",
         "Master call list (4x8 site owners + REC chairs). Set Result_Status as you call: "
         "To Call → Voicemail / No Answer / Reached - Pending → Approved / Declined."),
        ("Replacement_Bench",
         "When a primary 4x8 site declines, swap in the highest-scoring replacement "
         "from the same county."),
    ]
    for sheet, desc in workflow:
        ws.cell(row=r, column=1, value=sheet).font = h3_font
        ws.cell(row=r, column=2, value=desc).font = body
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 36
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="CALL SCRIPT (suggested)").font = h2_font
    r += 1
    script = (
        '"Hi [Owner First Name], this is [Caller] with the [Candidate] campaign — '
        'we\'re putting up 4×8 signs across [County] on properties with good road traffic, '
        'and yours at [Cross Street] is exactly the kind of spot we\'re looking for. '
        'We install it, we maintain it, we take it down after the election — '
        'would you be open to hosting one through November?"'
    )
    ws.cell(row=r, column=1, value=script).font = body
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 70
    r += 2

    ws.cell(row=r, column=1, value="STATUS DROPDOWN GUIDE").font = h2_font
    r += 1
    status_guide = [
        ("4x8_Sites — Install Status",  "Pending → Approved → Installed → Removed"),
        ("4x8_Sites — Approval Status", "Pending → Approved / Declined / Awaiting Permit"),
        ("Yard_Sign_RECs — Status",     "Scheduled → Confirmed → In Transit → Delivered"),
        ("Call_Queue — Result Status",  "To Call → Voicemail / No Answer / Reached - Pending → Approved / Declined"),
    ]
    for label, desc in status_guide:
        ws.cell(row=r, column=1, value=label).font = h3_font
        ws.cell(row=r, column=2, value=desc).font = body
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    widths = [38, 30, 18, 18, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: 4x8_Sites ──────────────────────────────────────────────────
    print(f"Writing 4x8_Sites ({n_primary} rows) ...")
    ws = wb.create_sheet("4x8_Sites")
    cols_4x8 = [
        "County", "Tier", "Rank", "Site_Name", "Address", "Maps_Link",
        "Phone", "Email", "Website",
        "Highway", "AADT_vehicles_per_day",
        "Suggested_Storefronts",
        "Approval_Status", "Install_Status", "Install_Date",
        "Owner_Contact", "Notes",
        "Candidate_ID",
    ]
    for ci, h in enumerate(cols_4x8, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = table_header_font; c.fill = table_header_fill; c.alignment = center

    # Sort by Tier (A first), County, then Composite_Score desc
    tier_order = {"A": 0, "B": 1, "C": 2, "D": 3}
    pp = prim.copy()
    pp["__tier_order__"] = pp["Tier"].map(tier_order).fillna(9)
    pp = pp.sort_values(["__tier_order__", "County", "Composite_Score"],
                        ascending=[True, True, False]).reset_index(drop=True)

    for idx, r in pp.iterrows():
        row = idx + 2
        ws.cell(row=row, column=1, value=_str(r.get("County")))
        ws.cell(row=row, column=2, value=_str(r.get("Tier")))
        ws.cell(row=row, column=3, value=int(r.get("Rank_in_County") or 0))
        ws.cell(row=row, column=4, value=_str(r.get("Name")))
        ws.cell(row=row, column=5, value=_str(r.get("Address")))
        # Hyperlink for Maps
        maps = _str(r.get("Maps_Link"))
        if maps:
            cell = ws.cell(row=row, column=6, value="View Map")
            cell.hyperlink = maps
            cell.font = Font(color="1F4E79", underline="single")
        ws.cell(row=row, column=7, value=_str(r.get("Phone")))
        ws.cell(row=row, column=8, value=_str(r.get("Email")))
        web = _str(r.get("Website"))
        if web and web.startswith("http"):
            cell = ws.cell(row=row, column=9, value=web)
            cell.hyperlink = web
            cell.font = Font(color="1F4E79", underline="single")
        else:
            ws.cell(row=row, column=9, value=web)
        ws.cell(row=row, column=10, value=_str(r.get("Nearest_Highway")))
        aadt = r.get("AADT")
        try:
            ws.cell(row=row, column=11, value=int(aadt) if aadt and aadt == aadt else None)
        except (TypeError, ValueError):
            ws.cell(row=row, column=11, value=None)
        # Suggested_Storefronts (newline-joined; up to 3)
        sf_parts = []
        for i in (1, 2, 3):
            sf_raw = r.get(f"Nearest_Storefront_{i}")
            if sf_raw is None or (isinstance(sf_raw, float) and sf_raw != sf_raw):
                continue
            sf = str(sf_raw).strip()
            if not sf or sf.lower() in ("nan", "none", "null"):
                continue
            d = r.get(f"Storefront_{i}_Distance_M")
            if d is None or (isinstance(d, float) and d != d):
                sf_parts.append(sf)
            else:
                try:
                    sf_parts.append(f"{sf} ({int(d)}m)")
                except (TypeError, ValueError):
                    sf_parts.append(sf)
        ws.cell(row=row, column=12, value="\n".join(sf_parts) if sf_parts else "").alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True)
        # Apply overlays from previous edits (idempotent preservation)
        cid = _str(r.get("Candidate_ID"))
        ov = sites_overlay.get(cid, {})
        ws.cell(row=row, column=13, value=ov.get("Approval_Status", "Pending"))
        ws.cell(row=row, column=14, value=ov.get("Install_Status", "Pending"))
        ws.cell(row=row, column=15, value=ov.get("Install_Date", ""))
        ws.cell(row=row, column=16, value=ov.get("Owner_Contact", ""))
        ws.cell(row=row, column=17, value=ov.get("Notes", ""))
        # Phone/Email overrides take precedence if user added one
        if "Phone" in ov:
            ws.cell(row=row, column=7, value=ov["Phone"])
        if "Email" in ov:
            ws.cell(row=row, column=8, value=ov["Email"])
        ws.cell(row=row, column=18, value=cid)

        # Color row by tier
        tier = _str(r.get("Tier")).upper()
        tier_fill = {"A": "FFF1F1", "B": "FFF8E8", "C": "EAF3FB", "D": "F4F4F4"}.get(tier)
        if tier_fill:
            for c in range(1, len(cols_4x8) + 1):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=tier_fill)

    # Header style
    ws.row_dimensions[1].height = 22
    n_rows = n_primary + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols_4x8))}{n_rows}"
    ws.freeze_panes = "A2"

    # Dropdowns — Approval_Status is now col M, Install_Status is col N
    install_dv = DataValidation(type="list",
        formula1='"Pending,Approved,Declined,Installed,Removed,Damaged,Site Issue"',
        allow_blank=True)
    install_dv.add(f"N2:N{n_rows}")
    ws.add_data_validation(install_dv)

    approval_dv = DataValidation(type="list",
        formula1='"Pending,Approved,Declined,Awaiting Permit,Permit Denied"',
        allow_blank=True)
    approval_dv.add(f"M2:M{n_rows}")
    ws.add_data_validation(approval_dv)

    widths_4x8 = [13, 6, 6, 32, 32, 11, 22, 26, 30, 12, 12, 36, 16, 16, 12, 22, 30, 12]
    for i, w in enumerate(widths_4x8, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Yard_Sign_RECs ─────────────────────────────────────────────
    print(f"Writing Yard_Sign_RECs ({len(ysa)} rows) ...")
    ws = wb.create_sheet("Yard_Sign_RECs")
    cols_yard = [
        "County", "Tier", "Plan_4000_Quantity", "Wave",
        "Chair_Name", "Chair_Phone", "Chair_Email",
        "REC_General_Email", "REC_Website",
        "Delivery_Contact", "Delivery_Phone", "Delivery_Email",
        "Drop_Address", "Region_Hub",
        "Status", "Delivery_Date", "Notes", "Followup_Required",
    ]
    for ci, h in enumerate(cols_yard, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = table_header_font; c.fill = table_header_fill; c.alignment = center

    yp = ysa.copy()
    yp["__tier_order__"] = yp["Strategic_Tier"].map(tier_order).fillna(9)
    yp = yp.sort_values(["__tier_order__", "Plan_4000"], ascending=[True, False]).reset_index(drop=True)

    for idx, r in yp.iterrows():
        row = idx + 2
        county = _str(r.get("County"))
        ov = recs_overlay.get(county, {})
        ws.cell(row=row, column=1, value=county)
        ws.cell(row=row, column=2, value=_str(r.get("Strategic_Tier")))
        ws.cell(row=row, column=3, value=int(r.get("Plan_4000") or 0))
        ws.cell(row=row, column=4, value=_str(r.get("Wave")))
        ws.cell(row=row, column=5, value=_str(r.get("Chair")))
        # Phone/Email overrides take precedence (field staff may have corrected)
        ws.cell(row=row, column=6, value=ov.get("Chair_Phone", _str(r.get("Chair_Phone"))))
        ws.cell(row=row, column=7, value=ov.get("Chair_Email", _str(r.get("Chair_Email"))))
        ws.cell(row=row, column=8, value=_str(r.get("REC_General_Email")))
        web = _str(r.get("REC_Website"))
        if web and web.startswith("http"):
            cell = ws.cell(row=row, column=9, value=web)
            cell.hyperlink = web
            cell.font = Font(color="1F4E79", underline="single")
        else:
            ws.cell(row=row, column=9, value=web)
        ws.cell(row=row, column=10, value=_str(r.get("Primary_Delivery_Contact")))
        ws.cell(row=row, column=11, value=_str(r.get("Primary_Delivery_Phone")))
        ws.cell(row=row, column=12, value=_str(r.get("Primary_Delivery_Email")))
        ws.cell(row=row, column=13, value=_str(r.get("Meeting_Delivery_Location")))
        ws.cell(row=row, column=14, value=_str(r.get("Region_Hub")))
        ws.cell(row=row, column=15, value=ov.get("Status", "Scheduled"))
        ws.cell(row=row, column=16, value=ov.get("Delivery_Date", ""))
        ws.cell(row=row, column=17, value=ov.get("Notes", ""))
        ws.cell(row=row, column=18, value=_str(r.get("Followup_Required")))

        # Color by POC status
        poc = _str(r.get("POC_Status_Yard"))
        poc_fill = {
            "chair_complete":          "D4EDDA",
            "partial_with_delivery":   "FFF3CD",
            "partial":                 "FFF3CD",
            "delivery_only":           "EAF3FB",
        }.get(poc)
        if poc_fill:
            for c in range(1, len(cols_yard) + 1):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=poc_fill)

    n_yard_rows = len(yp) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols_yard))}{n_yard_rows}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    yard_status_dv = DataValidation(type="list",
        formula1='"Scheduled,Confirmed,In Transit,Delivered,Issue,Cancelled"',
        allow_blank=True)
    yard_status_dv.add(f"O2:O{n_yard_rows}")
    ws.add_data_validation(yard_status_dv)

    widths_yard = [14, 6, 10, 10, 22, 16, 30, 30, 32, 22, 16, 30, 60, 28, 14, 14, 30, 40]
    for i, w in enumerate(widths_yard, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: Call_Queue ─────────────────────────────────────────────────
    print(f"Writing Call_Queue ({len(olog)} rows) ...")
    ws = wb.create_sheet("Call_Queue")
    cols_call = [
        "County", "Contact_Type", "Organization", "Contact_Name",
        "Phone_or_Email", "Method", "Result_Status",
        "Call_Date", "Outcome_Notes", "Next_Step", "Owner",
        "Source_ID",
    ]
    for ci, h in enumerate(cols_call, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = table_header_font; c.fill = table_header_fill; c.alignment = center

    # Sort: REC calls first (delivery is time-critical), then 4x8 by County
    cp = olog.copy()
    cp["__priority__"] = cp.get("Contact Type", "").apply(
        lambda v: 0 if "REC Chair" in str(v) else 1)
    cp = cp.sort_values(["__priority__", "County"]).reset_index(drop=True)

    for idx, r in cp.iterrows():
        row = idx + 2
        ws.cell(row=row, column=1, value=_str(r.get("County")))
        ws.cell(row=row, column=2, value=_str(r.get("Contact Type")))
        ws.cell(row=row, column=3, value=_str(r.get("Organization")))
        ws.cell(row=row, column=4, value=_str(r.get("Contact Name")))
        ws.cell(row=row, column=5, value=_str(r.get("Phone / Email Used")))
        ws.cell(row=row, column=6, value=_str(r.get("Method")))
        ws.cell(row=row, column=7, value=_str(r.get("Result Status")) or "To Call")
        ws.cell(row=row, column=8, value="")
        ws.cell(row=row, column=9, value="")
        ws.cell(row=row, column=10, value=_str(r.get("Next Step")))
        ws.cell(row=row, column=11, value="")
        ws.cell(row=row, column=12, value=_str(r.get("Source / Related Record")))

        # Color REC rows differently
        if "REC Chair" in _str(r.get("Contact Type")):
            for c in range(1, len(cols_call) + 1):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="FFF1F1")

    n_call_rows = len(cp) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols_call))}{n_call_rows}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    call_dv = DataValidation(type="list",
        formula1='"To Call,Voicemail,No Answer,Reached - Pending,Approved,Declined,Wrong Number,Do Not Contact,Closed"',
        allow_blank=True)
    call_dv.add(f"G2:G{n_call_rows}")
    ws.add_data_validation(call_dv)

    widths_call = [14, 18, 36, 24, 32, 14, 18, 12, 40, 22, 14, 14]
    for i, w in enumerate(widths_call, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 5: Replacement_Bench ──────────────────────────────────────────
    print(f"Writing Replacement_Bench ({n_repl} rows) ...")
    ws = wb.create_sheet("Replacement_Bench")
    cols_rep = [
        "County", "Tier", "Site_Name", "Address", "Maps_Link",
        "Phone", "Email", "Website",
        "Highway", "AADT_vehicles_per_day",
        "Composite_Score", "Candidate_ID",
    ]
    for ci, h in enumerate(cols_rep, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = table_header_font; c.fill = table_header_fill; c.alignment = center

    rp = repl.copy()
    rp["__tier_order__"] = rp["Tier"].map(tier_order).fillna(9)
    rp = rp.sort_values(["__tier_order__", "County", "Composite_Score"],
                        ascending=[True, True, False]).reset_index(drop=True)
    for idx, r in rp.iterrows():
        row = idx + 2
        ws.cell(row=row, column=1, value=_str(r.get("County")))
        ws.cell(row=row, column=2, value=_str(r.get("Tier")))
        ws.cell(row=row, column=3, value=_str(r.get("Name")))
        ws.cell(row=row, column=4, value=_str(r.get("Address")))
        maps = _str(r.get("Maps_Link"))
        if maps:
            cell = ws.cell(row=row, column=5, value="View Map")
            cell.hyperlink = maps
            cell.font = Font(color="1F4E79", underline="single")
        ws.cell(row=row, column=6, value=_str(r.get("Phone")))
        ws.cell(row=row, column=7, value=_str(r.get("Email")))
        ws.cell(row=row, column=8, value=_str(r.get("Website")))
        ws.cell(row=row, column=9, value=_str(r.get("Nearest_Highway")))
        try:
            ws.cell(row=row, column=10,
                    value=int(r.get("AADT")) if r.get("AADT") and r.get("AADT") == r.get("AADT") else None)
        except (TypeError, ValueError):
            ws.cell(row=row, column=10, value=None)
        try:
            ws.cell(row=row, column=11, value=round(float(r.get("Composite_Score")), 1))
        except (TypeError, ValueError):
            ws.cell(row=row, column=11, value=None)
        ws.cell(row=row, column=12, value=_str(r.get("Candidate_ID")))

    n_rep_rows = n_repl + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols_rep))}{n_rep_rows}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    widths_rep = [14, 6, 38, 36, 11, 22, 28, 32, 14, 14, 14, 12]
    for i, w in enumerate(widths_rep, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Save
    wb.save(out)
    size_kb = out.stat().st_size / 1024
    print(f"\nWrote {out} ({size_kb:.0f} KB)")
    print(f"  5 sheets: Start_Here | 4x8_Sites ({n_primary}) | Yard_Sign_RECs ({len(ysa)}) | Call_Queue ({len(olog)}) | Replacement_Bench ({n_repl})")
    return 0


def _str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v != v:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "null") else s


if __name__ == "__main__":
    sys.exit(main())
